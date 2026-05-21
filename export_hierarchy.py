import argparse
from concurrent.futures import ThreadPoolExecutor, as_completed
import csv
import json
import logging
import re
import sys
import threading
import time
from pathlib import Path
from urllib.parse import parse_qsl, urlencode, urlparse, urlsplit, urlunsplit

from openpyxl import Workbook
import requests

# API Endpoints
AUTH_URL = "https://siemens-bt-015.eu.auth0.com/oauth/token"
BASE_API_URL = "https://api.bpcloud.siemens.com/operations"
AUDIENCE = "https://horizon.siemens.com"
OUTPUT_DIR = Path("output")
OUTPUT_FILE = OUTPUT_DIR / "Building_Operations_Data.xlsx"
LOG_FILE = OUTPUT_DIR / "export_hierarchy.csv"
PAGE_LIMIT = 100
MAX_PAGE_LIMIT = 100
DEFAULT_TIMEOUT_SECONDS = 30
DEFAULT_RATE_LIMIT_PER_SECOND = 8
DEFAULT_MAX_WORKERS = 4
LOCATION_COLUMNS = ["Location ID", "Location Name"]
DEVICE_COLUMNS = [
    "Device ID",
    "Device Name",
    "Device Description",
    "Location ID",
    "Location Name",
    "Parent Device ID",
    "Parent Device Name",
]
POINT_COLUMNS = [
    "Point ID",
    "Point Name",
    "Point Description",
    "Point Is Active",
    "Device ID",
    "Device Name",
    "Location ID",
    "Location Name",
]


class CsvLogFormatter(logging.Formatter):
    def format(self, record):
        message = record.getMessage()
        if record.exc_info:
            message = f"{message}\n{self.formatException(record.exc_info)}"

        output = CsvListWriter()
        writer = csv.writer(output, lineterminator="")
        writer.writerow(
            [
                self.formatTime(record, "%Y-%m-%d %H:%M:%S"),
                record.levelname,
                record.name,
                message,
            ]
        )
        return output.value


class CsvListWriter:
    def __init__(self):
        self.value = ""

    def write(self, value):
        self.value += value


def configure_logging(debug=False):
    level = logging.DEBUG if debug else logging.INFO
    OUTPUT_DIR.mkdir(exist_ok=True)
    formatter = logging.Formatter("%(asctime)s %(levelname)s: %(message)s", "%H:%M:%S")
    root_logger = logging.getLogger()
    root_logger.setLevel(level)
    root_logger.handlers.clear()

    console_handler = logging.StreamHandler()
    console_handler.setLevel(level)
    console_handler.setFormatter(formatter)
    root_logger.addHandler(console_handler)

    with LOG_FILE.open("w", encoding="utf-8", newline="") as f:
        csv.writer(f).writerow(["timestamp", "level", "logger", "message"])

    file_handler = logging.FileHandler(LOG_FILE, mode="a", encoding="utf-8")
    file_handler.setLevel(logging.DEBUG)
    file_handler.setFormatter(CsvLogFormatter())
    root_logger.addHandler(file_handler)

    logging.getLogger("urllib3").setLevel(logging.WARNING)
    logging.getLogger("charset_normalizer").setLevel(logging.WARNING)


def load_auth_secrets():
    secrets_path = Path(__file__).with_name("secrets.json")
    with secrets_path.open("r", encoding="utf-8") as f:
        secrets = json.load(f)

    client = secrets.get("client", {})
    client_id = client.get("client_id")
    client_secret = client.get("client_secret")
    partition_id = client.get("partition_id")

    missing = [
        name
        for name, value in (
            ("client.client_id", client_id),
            ("client.client_secret", client_secret),
            ("client.partition_id", partition_id),
        )
        if not value
    ]
    if missing:
        raise ValueError(f"Missing required secrets.json value(s): {', '.join(missing)}")

    return client_id, client_secret, partition_id


CLIENT_ID = None
CLIENT_SECRET = None
PARTITION_ID = None
RATE_LIMITER = None
THREAD_LOCAL = threading.local()


class RateLimiter:
    def __init__(self, requests_per_second):
        self.min_interval = 1 / requests_per_second
        self.lock = threading.Lock()
        self.next_request_at = 0

    def wait(self):
        with self.lock:
            now = time.monotonic()
            wait_seconds = self.next_request_at - now
            if wait_seconds > 0:
                time.sleep(wait_seconds)
                now = time.monotonic()
            self.next_request_at = now + self.min_interval


def get_http_session():
    session = getattr(THREAD_LOCAL, "session", None)
    if session is None:
        session = requests.Session()
        THREAD_LOCAL.session = session
    return session


def request_json(method, url, *, timeout, **kwargs):
    logging.debug("%s %s", method.upper(), url)
    try:
        if RATE_LIMITER:
            RATE_LIMITER.wait()
        response = get_http_session().request(method, url, timeout=timeout, **kwargs)
        response.raise_for_status()
        try:
            return json.loads(response.text)
        except json.JSONDecodeError as exc:
            repaired_text = repair_malformed_json(response.text)
            if repaired_text != response.text:
                try:
                    logging.warning(
                        "Invalid JSON repaired for %s %s. %s",
                        method.upper(),
                        url,
                        format_json_error(response.text, exc),
                    )
                    return json.loads(repaired_text)
                except json.JSONDecodeError:
                    pass

            retry_url = remove_query_params(url, {"include"})
            if retry_url != url:
                logging.warning(
                    "Invalid JSON from included resources; retrying without include parameter. %s",
                    format_json_error(response.text, exc),
                )
                return request_json(method, retry_url, timeout=timeout, **kwargs)
            raise RuntimeError(
                f"Invalid JSON from {method.upper()} {url}: {format_json_error(response.text, exc)}"
            ) from exc
    except requests.Timeout as exc:
        raise RuntimeError(f"Request timed out after {timeout}s: {method.upper()} {url}") from exc
    except requests.HTTPError as exc:
        body = response.text[:1000]
        raise RuntimeError(
            f"HTTP {response.status_code} from {method.upper()} {url}: {body}"
        ) from exc
    except requests.RequestException as exc:
        raise RuntimeError(f"Request failed for {method.upper()} {url}: {exc}") from exc


def format_json_error(text, exc):
    start = max(0, exc.pos - 300)
    end = min(len(text), exc.pos + 300)
    snippet = text[start:end].replace("\r", "\\r").replace("\n", "\\n")
    return (
        f"{exc.msg} at line {exc.lineno}, column {exc.colno}, char {exc.pos}. "
        f"Response snippet near error: {snippet}"
    )


def repair_malformed_json(text):
    """Best-effort repair for API string fields with invalid escaping/quotes."""
    result = []
    in_string = False
    escaped = False
    index = 0

    while index < len(text):
        char = text[index]

        if not in_string:
            result.append(char)
            if char == '"':
                in_string = True
            index += 1
            continue

        if escaped:
            if char in '"\\/bfnrt':
                result.append("\\" + char)
            elif char == "u" and is_unicode_escape(text, index):
                result.append("\\" + text[index : index + 5])
                index += 4
            else:
                result.append("\\\\" + char)
            escaped = False
            index += 1
            continue

        if char == "\\":
            escaped = True
            index += 1
            continue

        if char == '"':
            if quote_looks_like_string_end(text, index):
                result.append(char)
                in_string = False
            else:
                result.append('\\"')
            index += 1
            continue

        codepoint = ord(char)
        if codepoint < 0x20:
            result.append(escape_control_character(char))
        else:
            result.append(char)
        index += 1

    if escaped:
        result.append("\\\\")

    return "".join(result)


def is_unicode_escape(text, index):
    return index + 4 < len(text) and re.fullmatch(r"[0-9a-fA-F]{4}", text[index + 1 : index + 5])


def quote_looks_like_string_end(text, quote_index):
    index = quote_index + 1
    while index < len(text) and text[index].isspace():
        index += 1
    return index == len(text) or text[index] in ",:}]"


def escape_control_character(char):
    if char == "\n":
        return "\\n"
    if char == "\r":
        return "\\r"
    if char == "\t":
        return "\\t"
    return f"\\u{ord(char):04x}"


def remove_query_params(url, params_to_remove):
    parts = urlsplit(url)
    query = [
        (key, value)
        for key, value in parse_qsl(parts.query, keep_blank_values=True)
        if key not in params_to_remove
    ]
    return urlunsplit((parts.scheme, parts.netloc, parts.path, urlencode(query), parts.fragment))


def get_query_param(url, param):
    parts = urlsplit(url)
    for key, value in parse_qsl(parts.query, keep_blank_values=True):
        if key == param:
            return value
    return None


def set_query_param(url, param, value):
    parts = urlsplit(url)
    query = [
        (key, existing_value)
        for key, existing_value in parse_qsl(parts.query, keep_blank_values=True)
        if key != param
    ]
    query.append((param, value))
    return urlunsplit((parts.scheme, parts.netloc, parts.path, urlencode(query), parts.fragment))


def paged_get(url, headers, *, timeout):
    """Yield items from a JSON:API collection, following links.next when present."""
    next_url = url
    include_value = get_query_param(url, "include")
    page = 1

    while next_url:
        result = request_json("get", next_url, headers=headers, timeout=timeout)
        data = result.get("data", [])
        logging.debug("Fetched page %s with %s item(s)", page, len(data))
        yield result

        next_link = result.get("links", {}).get("next")
        next_url = build_next_url(next_link)
        if include_value and next_url and not get_query_param(next_url, "include"):
            next_url = set_query_param(next_url, "include", include_value)
        page += 1


def build_next_url(next_link):
    if not next_link:
        return None
    if next_link.startswith("http"):
        return next_link

    parsed_base = urlparse(BASE_API_URL)
    origin = f"{parsed_base.scheme}://{parsed_base.netloc}"
    if next_link.startswith("/operations/"):
        return f"{origin}{next_link}"
    if next_link.startswith("/"):
        return f"{BASE_API_URL}{next_link}"
    return f"{BASE_API_URL}/{next_link}"


def first_nonempty(*values, default="N/A"):
    for value in values:
        if value is not None and str(value).strip():
            return value
    return default


def format_duration(seconds):
    seconds = max(0, int(seconds))
    hours, remainder = divmod(seconds, 3600)
    minutes, seconds = divmod(remainder, 60)
    if hours:
        return f"{hours}h {minutes}m {seconds}s"
    if minutes:
        return f"{minutes}m {seconds}s"
    return f"{seconds}s"


def collect_device_info(page):
    device_info_by_id = {}
    for item in page.get("included", []):
        if item.get("type") != "DeviceInfo":
            continue
        device_id = (
            item.get("relationships", {})
            .get("hasDevice", {})
            .get("data", {})
            .get("id")
        )
        if device_id:
            device_info_by_id[device_id] = item.get("attributes", {})
    return device_info_by_id


def build_device_row(dev, device_info, location_id, parent_device=None):
    dev_id = dev.get("id")
    dev_attributes = dev.get("attributes", {})
    dev_profile = dev_attributes.get("profile", {})
    dev_name = first_nonempty(
        device_info.get("name"),
        dev_profile.get("name"),
        dev_attributes.get("name"),
        dev_attributes.get("label"),
        dev_id,
    )
    dev_description = first_nonempty(
        device_info.get("description"),
        dev_attributes.get("description"),
        default="No Description",
    )

    return {
        "Device ID": dev_id,
        "Device Name": dev_name,
        "Device Description": dev_description,
        "Location ID": location_id,
        "Parent Device ID": parent_device.get("Device ID") if parent_device else "",
        "Parent Device Name": parent_device.get("Device Name") if parent_device else "",
    }


def get_access_token(timeout):
    """Retrieve an OAuth2 Bearer token."""
    logging.info("Requesting access token")
    payload = {
        "client_id": CLIENT_ID,
        "client_secret": CLIENT_SECRET,
        "audience": AUDIENCE,
        "grant_type": "client_credentials",
    }
    headers = {"Content-Type": "application/json"}
    result = request_json("post", AUTH_URL, json=payload, headers=headers, timeout=timeout)
    token = result.get("access_token")
    if not token:
        raise RuntimeError("Token response did not include access_token")
    return token


def fetch_locations(token, timeout):
    """Fetch all locations."""
    logging.info("Fetching building locations")
    url = (
        f"{BASE_API_URL}/partitions/{PARTITION_ID}/locations"
        "?filter[type]=Building&fields[Building]=label"
    )
    headers = {"Authorization": f"Bearer {token}"}
    result = request_json("get", url, headers=headers, timeout=timeout)
    locations = []
    for loc in result.get("data", []):
        loc_id = loc.get("id")
        loc_name = loc.get("attributes", {}).get("label", "N/A")
        if not loc_id:
            logging.warning("Skipping location without an id: %s", loc)
            continue
        logging.debug("Found location: %s (%s)", loc_name, loc_id)
        locations.append({"Location ID": loc_id, "Location Name": loc_name})
    logging.info("Found %s building location(s)", len(locations))
    return locations


def fetch_devices_for_location(token, location_id, timeout):
    """Fetch all devices for a given location."""
    url = (
        f"{BASE_API_URL}/partitions/{PARTITION_ID}/devices"
        f"?filter[hasLocation.data.id]={location_id}"
        f"&page[limit]={PAGE_LIMIT}&include=hasFeatures.DeviceInfo"
    )
    headers = {"Authorization": f"Bearer {token}"}
    devices_data = []

    logging.debug("Fetching devices for location %s", location_id)
    try:
        pages = paged_get(url, headers, timeout=timeout)
        for page in pages:
            devices = page.get("data", [])
            device_info_by_id = collect_device_info(page)

            for dev in devices:
                dev_id = dev.get("id")
                if not dev_id:
                    logging.warning("Skipping device without an id: %s", dev)
                    continue

                devices_data.append(build_device_row(dev, device_info_by_id.get(dev_id, {}), location_id))
    except RuntimeError as exc:
        logging.error("Skipping remaining device pages for location %s: %s", location_id, exc)

    logging.info("Found %s device(s) for location %s", len(devices_data), location_id)
    return devices_data


def fetch_devices_behind_device(token, parent_device, timeout):
    """Fetch devices behind a gateway or parent device."""
    parent_id = parent_device.get("Device ID")
    location_id = parent_device.get("Location ID")
    url = (
        f"{BASE_API_URL}/partitions/{PARTITION_ID}/devices/{parent_id}/devices"
        f"?page[limit]={PAGE_LIMIT}&include=hasFeatures.DeviceInfo"
    )
    headers = {"Authorization": f"Bearer {token}"}
    devices_data = []

    logging.debug("Fetching child devices behind %s", parent_id)
    try:
        pages = paged_get(url, headers, timeout=timeout)
        for page in pages:
            devices = page.get("data", [])
            device_info_by_id = collect_device_info(page)

            for dev in devices:
                dev_id = dev.get("id")
                if not dev_id:
                    logging.warning("Skipping child device without an id: %s", dev)
                    continue
                devices_data.append(
                    build_device_row(
                        dev,
                        device_info_by_id.get(dev_id, {}),
                        location_id,
                        parent_device=parent_device,
                    )
                )
    except RuntimeError as exc:
        if "HTTP 404" in str(exc):
            logging.debug("No child-device endpoint data for %s", parent_id)
            return []
        raise

    if devices_data:
        logging.info("Found %s child device(s) behind %s", len(devices_data), parent_id)
    return devices_data


def fetch_all_devices_for_location(token, location_id, timeout, include_gateway_devices=True):
    """Fetch devices assigned to a location, optionally walking gateway child devices."""
    root_devices = fetch_devices_for_location(token, location_id, timeout)
    seen_device_ids = set()
    all_devices = []
    queue = []

    for device in root_devices:
        device_id = device.get("Device ID")
        if device_id in seen_device_ids:
            continue
        seen_device_ids.add(device_id)
        all_devices.append(device)
        queue.append(device)

    if not include_gateway_devices:
        return all_devices

    while queue:
        parent_device = queue.pop(0)
        child_devices = fetch_devices_behind_device(token, parent_device, timeout)
        for child in child_devices:
            child_id = child.get("Device ID")
            if child_id in seen_device_ids:
                continue
            seen_device_ids.add(child_id)
            all_devices.append(child)
            queue.append(child)

    logging.info("Found %s total unique device(s) for location %s", len(all_devices), location_id)
    return all_devices


def iter_points_for_device(token, device_id, device_name, timeout):
    """Yield points for a given device as API pages are fetched."""
    url = (
        f"{BASE_API_URL}/partitions/{PARTITION_ID}/devices/{device_id}/points"
        f"?page[limit]={PAGE_LIMIT}"
    )
    headers = {"Authorization": f"Bearer {token}"}
    logging.debug("Fetching points for device %s (%s)", device_name, device_id)
    for page in paged_get(url, headers, timeout=timeout):
        for pt in page.get("data", []):
            attributes = pt.get("attributes", {})
            yield {
                "Point ID": pt.get("id"),
                "Point Name": attributes.get("name", "N/A"),
                "Point Description": attributes.get("description", ""),
                "Point Is Active": attributes.get("isActive", ""),
                "Device ID": device_id,
                "Device Name": device_name,
            }


def fetch_points_for_device(token, device_id, device_name, timeout):
    """Fetch all points for a given device."""
    points_data = list(iter_points_for_device(token, device_id, device_name, timeout))
    logging.info("Found %s point(s) for device %s", len(points_data), device_id)
    return points_data


def append_row(sheet, row, columns):
    sheet.append([row.get(column, "") for column in columns])


def fetch_point_rows_for_device(token, device_row, timeout):
    dev_id = device_row.get("Device ID")
    dev_name = device_row.get("Device Name")
    point_rows = []

    try:
        for pt in iter_points_for_device(token, dev_id, dev_name, timeout):
            point_rows.append(
                {
                "Point ID": pt.get("Point ID"),
                "Point Name": pt.get("Point Name"),
                "Point Description": pt.get("Point Description"),
                "Point Is Active": pt.get("Point Is Active"),
                "Device ID": dev_id,
                "Device Name": dev_name,
                    "Location ID": device_row.get("Location ID"),
                    "Location Name": device_row.get("Location Name"),
                }
            )
    except RuntimeError as exc:
        logging.error("Skipping points for device %s (%s): %s", dev_name, dev_id, exc)

    return device_row, point_rows


def parse_args():
    parser = argparse.ArgumentParser(
        description="Export Building X locations, devices, and points to an Excel workbook."
    )
    parser.add_argument(
        "--debug",
        action="store_true",
        help="Show each request and page fetch while the script runs.",
    )
    parser.add_argument(
        "--timeout",
        type=int,
        default=DEFAULT_TIMEOUT_SECONDS,
        help=f"HTTP request timeout in seconds. Defaults to {DEFAULT_TIMEOUT_SECONDS}.",
    )
    parser.add_argument(
        "--page-limit",
        type=int,
        default=PAGE_LIMIT,
        help=f"Records to request per API page. Defaults to {PAGE_LIMIT}; max {MAX_PAGE_LIMIT}.",
    )
    parser.add_argument(
        "--rate-limit",
        type=float,
        default=DEFAULT_RATE_LIMIT_PER_SECOND,
        help=(
            "Maximum API requests per second across all workers. "
            f"Defaults to {DEFAULT_RATE_LIMIT_PER_SECOND}."
        ),
    )
    parser.add_argument(
        "--max-workers",
        type=int,
        default=DEFAULT_MAX_WORKERS,
        help=f"Concurrent devices to fetch points for. Defaults to {DEFAULT_MAX_WORKERS}.",
    )
    parser.add_argument(
        "--include-gateway-devices",
        action="store_true",
        help="Also walk devices behind each gateway, de-duplicating repeated devices.",
    )
    return parser.parse_args()


def main():
    global CLIENT_ID, CLIENT_SECRET, PARTITION_ID, PAGE_LIMIT, RATE_LIMITER

    args = parse_args()
    configure_logging(args.debug)
    if args.page_limit < 1 or args.page_limit > MAX_PAGE_LIMIT:
        raise ValueError(f"--page-limit must be between 1 and {MAX_PAGE_LIMIT}")
    if args.rate_limit <= 0 or args.rate_limit >= 10:
        raise ValueError("--rate-limit must be greater than 0 and less than 10")
    if args.max_workers < 1:
        raise ValueError("--max-workers must be at least 1")
    PAGE_LIMIT = args.page_limit
    RATE_LIMITER = RateLimiter(args.rate_limit)

    logging.info("Starting hierarchy export")
    logging.info("Using API page limit: %s", PAGE_LIMIT)
    logging.info(
        "Using up to %s point worker(s) with %.2f request(s)/second limit",
        args.max_workers,
        args.rate_limit,
    )
    export_started_at = time.monotonic()
    CLIENT_ID, CLIENT_SECRET, PARTITION_ID = load_auth_secrets()
    token = get_access_token(args.timeout)

    OUTPUT_DIR.mkdir(exist_ok=True)
    workbook = Workbook(write_only=True)
    saved = False
    try:
        locations_sheet = workbook.create_sheet("Locations")
        devices_sheet = workbook.create_sheet("Devices")
        points_sheet = workbook.create_sheet("Points")
        locations_sheet.append(LOCATION_COLUMNS)
        devices_sheet.append(DEVICE_COLUMNS)
        points_sheet.append(POINT_COLUMNS)

        locations = fetch_locations(token, args.timeout)
        device_rows_for_points = []
        for loc in locations:
            loc_id = loc.get("Location ID")
            loc_name = loc.get("Location Name")
            logging.info("Processing location: %s (%s)", loc_name, loc_id)
            append_row(locations_sheet, loc, LOCATION_COLUMNS)

            devices = fetch_all_devices_for_location(
                token,
                loc_id,
                args.timeout,
                include_gateway_devices=args.include_gateway_devices,
            )
            for dev in devices:
                dev_id = dev.get("Device ID")
                dev_name = dev.get("Device Name")
                device_row = {
                    "Device ID": dev_id,
                    "Device Name": dev_name,
                    "Device Description": dev.get("Device Description"),
                    "Location ID": loc_id,
                    "Location Name": loc_name,
                    "Parent Device ID": dev.get("Parent Device ID"),
                    "Parent Device Name": dev.get("Parent Device Name"),
                }
                append_row(devices_sheet, device_row, DEVICE_COLUMNS)
                device_rows_for_points.append(device_row)

        discovery_elapsed = time.monotonic() - export_started_at
        total_devices = len(device_rows_for_points)
        logging.info(
            "Discovered %s location(s) and %s device(s) in %s. Starting points export.",
            len(locations),
            total_devices,
            format_duration(discovery_elapsed),
        )

        points_started_at = time.monotonic()
        total_points = 0
        max_workers = min(args.max_workers, total_devices) if total_devices else 1
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            futures = [
                executor.submit(fetch_point_rows_for_device, token, device_row, args.timeout)
                for device_row in device_rows_for_points
            ]

            for index, future in enumerate(as_completed(futures), start=1):
                device_row, point_rows = future.result()
                dev_id = device_row.get("Device ID")
                total_points += len(point_rows)
                for point_row in point_rows:
                    append_row(points_sheet, point_row, POINT_COLUMNS)
                logging.info("Found %s point(s) for device %s", len(point_rows), dev_id)

                elapsed = time.monotonic() - points_started_at
                average_seconds_per_device = elapsed / index
                remaining_devices = total_devices - index
                estimated_remaining = average_seconds_per_device * remaining_devices
                logging.info(
                    "Progress: %s/%s devices, %s point(s), elapsed %s, ETA %s",
                    index,
                    total_devices,
                    total_points,
                    format_duration(elapsed),
                    format_duration(estimated_remaining),
                )

        logging.info("Writing workbook: %s", OUTPUT_FILE)
        workbook.save(OUTPUT_FILE)
        saved = True

        total_elapsed = time.monotonic() - export_started_at
        logging.info(
            "Data has been successfully written to '%s' in %s.",
            OUTPUT_FILE,
            format_duration(total_elapsed),
        )
    finally:
        if not saved:
            workbook.close()


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        logging.critical("Export failed: %s", exc)
        sys.exit(1)
